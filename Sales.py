from dialog import getfiledir
from easygui import *
import datetime
import string
from Create_Dict_script import createdict
from openpyxl import load_workbook


##Written by Cameron Burt
##The goal of this script is to write out to a template that has been redacted for Gennaro Security Concerns
##This was part of a weekly reporting process for Gennaro.


#create list of customers
def createlistofcust(CustomerCode,Dictlookup):

    for i in range(len(CustomerCode)):
        CustomerCode[i] = Dictlookup.get(CustomerCode[i])
    customerfinallist = []
    i = 0
    while i<len(CustomerCode):
        if CustomerCode[i] not in customerfinallist:
            customerfinallist.append(CustomerCode[i])
            i=i+1
        else:
            i = i+1

    return customerfinallist
            
    
        
#Grab the data from a CSV. Format of input file =
#CustomerCode, Date, CustomerRef, Total, FiscalMonth (Int!)
def getdata(): 
    
    goodFile = False
    while goodFile == False:
        #calls to Dialog.py and opens a dialog box for the user to select a file. 
        fname = getfiledir()
        # Begin exception handling
        try:
            # Try to open the file using the name given
            olympicsFile = open(fname, 'r')
            # If the name is valid, set Boolean to true to exit loop
            goodFile = True
        except:
            # If the name is not valid - IOError exception is raised
            print("Invalid filename, please try again ... ")


    customerCode = []
    InvoiceTotal = []
    Date = []
    Ref = []
    FiscalYear = []
    FiscalMonth = []
    next(olympicsFile)
    #loop through file line by line and write to lists, sku, item, qty are the temp variables, SKU, Qty, Item are the lists. 
    for line in olympicsFile:
        line = line.strip()
        code, tempdate, Ref, Total, Month = line.split(',')
        customerCode.append(code)
        InvoiceTotal.append(float(Total))
        FiscalMonth.append(int(Month))

        
    olympicsFile.close

    return customerCode, InvoiceTotal, FiscalMonth




#Tally will loop through all data looking for the key, (customer code, and month), tally them all up, then send it back.
def tally(key, month, customerCode, InvoiceTotal, FiscalMonth):
    total= 0
    #loop through customer code and at each match of key check Fiscal Month,  If fiscal month is month, add to tally,
    for i in range(len(customerCode)):
        if (customerCode[i] == key): 
                if FiscalMonth[i] == month:
                    total = total+InvoiceTotal[i]

    
    return total



    
   
#creating a customer 2 customer name lookup
def createdict():

    ##GET DATA 
    goodFile = False
    while goodFile == False:
        #calls to Dialog.py and opens a dialog box for the user to select a file. 
        fname = getfiledir()
        # Begin exception handling
        try:
            # Try to open the file using the name given
            olympicsFile = open(fname, 'r')
            # If the name is valid, set Boolean to true to exit loop
            goodFile = True
        except:
            # If the name is not valid - IOError exception is raised
            print("Invalid filename, please try again ... ")


    List1 = []
    List2 = []
    #loop through file line by line and write to lists, sku, item, qty are the temp variables, SKU, Qty, Item are the lists. 
    for line in olympicsFile:
        line = line.strip()
        x, y = line.split(',')
        List1.append(x)
        List2.append(y)

        
    olympicsFile.close

    ##CREATE DICT STARTS
    lookupObj = zip(List1, List2)
    Dict1 = dict(lookupObj)
   
    


    return Dict1

    

def createsaleslists(salesperson, master):

    #Create 4 lists
    Ani = []
    Jord = []
    Lynds = []
    Other = []
    i = 0
    while i<len(master):
        if (salesperson.get(master[i][0]) == 'Salesperson1'):
            Ani.append(master[i])
        if (salesperson.get(master[i][0]) == 'Salesperson2'):
            Jord.append(master[i])
        if (salesperson.get(master[i][0]) == 'Salesperson3'):
            Lynds.append(master[i])
        if (salesperson.get(master[i][0]) == 'Salesperson4'):
            Other.append(master[i])
        i = i+1
            

    return Ani, Jord, Lynds, Other 




def insertionSort(arr): 
  
    # Traverse through 1 to len(arr) 
    for i in range(1, len(arr)): 
  
        key = arr[i] 
  
        # Move elements of arr[0..i-1], that are 
        # greater than key, to one position ahead 
        # of their current position 
        j = i-1
        while j >=0 and key < arr[j] : 
                arr[j+1] = arr[j] 
                j -= 1
        arr[j+1] = key 

    return arr
  

    

def main():
#Get information from the data file
    msgbox('Please choose the file which holds the Raw Sales Data')
    customerCode, InvoiceTotal, FiscalMonth = getdata()
    msgbox('Please choose the file which holds the customer lookup table')
    customer2namelookup = createdict()
    msgbox('Please choose the file which holds the salesperson lookup table')
    salesperson = createdict()
    CustList = createlistofcust(customerCode, customer2namelookup)

   # print(CustList)
    #print (len(CustList))


##    print (customerCode)
##    for i in range(len(customerCode)):
##        customerCode[i] = customer2namelookup.get(customerCode[i])

    #msgbox('Enter the target location')
    wb = load_workbook(filename = 'Weeklysalestemplate.xlsx')
    sheet = wb.active
    
    i = 0

    
    #for each customer, send it to tally.
    results = []
    while i<(len(CustList)):
        cust = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        cust[0] = CustList[i]
        j =1
        while (j<13):
            total = tally(CustList[i], j, customerCode , InvoiceTotal, FiscalMonth)
            cust[j] = total
            j = j + 1

    #LIST OF LISTS
        results.append(cust)
        i = i +1

        
    #send results to another func to write the salesperson lookup lists. 
    Ani, Jord, Lynds, Other = createsaleslists(salesperson, results)

    #sort Ani
    Ani = insertionSort(Ani)


    ###WRITE OUT
        #Idex: (2,3) is start of first customer (Ambisionista box)
    x = 3
    y = 3
        #format to write to a cell sheet.cell(row = x, column = y).value = stuff
        
    i = 0
    j = 1
        #ani loop from (3,4)- (3,15) for first customer
        #last customer on line (21,3) to (21,14)

    #Loops and writes to specific spots on table. Changes are not made constantly the excel file is fairly robust, as such, it is more convenient in this use case to
    #manually adjust the (X,Y) cell coordinates via the loop end numbers. If the excel sheet was more basic, and more dev time was allowed, one could feasibly change it dynamically without code edits.
    print ("Still working... Building Salesperson 1's Table")
    print (Ani)
    while (x <=23):
        y = 4
        j = 1
        #print ("Loop1")
        while (y <= 15):
            #print(Ani[i][j]+" Current Value")
            sheet.cell(row = x, column = y).value = Ani[i][j]
            j = j +1
            y = y + 1
        x = x + 1
        i = i + 1

    Jord = insertionSort(Jord)
    # WRITE OUT FOR JORDAN
    i = 0
    x = 26
    y = 3
    print ("Still working.. Building Salesperson 2's Table")
    print (Jord)
    while (x <=35):
        y = 4
        j = 1
        #print ("Loop1")
        while (y <= 15):
            sheet.cell(row = x, column = y).value = Jord[i][j]
            j = j +1
            y = y + 1
        x = x + 1
        i = i + 1
    #sort for lindsay
    Lynds = insertionSort(Lynds)
    #print (Lynds)

    ##Write out for lindsay
    i = 0
    x = 38
    y = 3
    print ("Still Working... Building Salesperson 3's Table")
    print (Lynds)
    while (x <=49):
        y = 4
        j = 1
        #print ("Loop1")
        while (y <= 15):
            sheet.cell(row = x, column = y).value = Lynds[i][j]
            j = j +1
            y = y + 1
        x = x + 1
        i = i + 1
    Other = insertionSort(Other)
    #print (Other)
    print ("Still Working... Handling Everything Else")
    print (Other)
    i=0
    x = 52
    y = 3
    while (x <=52):
        y = 4
        j = 1
        #print ("Loop1")
        while (y <= 15):
            sheet.cell(row = x, column = y).value = Other[i][j]
            j = j +1
            y = y + 1
        x = x + 1
        i = i + 1
    

    ##Save As Template with Date
    now = datetime.datetime.now()
    year= now.year
    month = now.month
    day = now.day
    date = str(year) + " - " + str(month) + " - " + str(day)
    savename = "Sales Report"+date+".xlsx"
    
    wb.defined_names.definedName = []
    wb.save(savename)
    
main()
